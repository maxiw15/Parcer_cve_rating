import requests
from bs4 import BeautifulSoup
import openpyxl


# Функция для получения информации о уязвимости по номеру CVE
def get_vulnerability_info(cve_number):
    url = f"https://nvd.nist.gov/vuln/detail/{cve_number}"

    try:
        # Отправляем запрос на сайт
        response = requests.get(url)
        response.raise_for_status()  # Проверяем наличие ошибок при запросе

        # Парсим HTML-страницу с помощью BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')

        # Находим элемент с информацией о CVSS
        cvss_element = soup.find('a', {'id': 'Cvss3NistCalculatorAnchor'})
        if cvss_element:
            cvss_score = cvss_element.text.strip()  # Получаем текст элемента
            return cvss_score
        else:
            return "CVSS информация не найдена"
    except Exception as e:
        return f"Ошибка при запросе: {e}"


# Открываем Excel-файл
workbook = openpyxl.load_workbook('CVE.xlsx')
sheet = workbook.active

# Проходимся по каждой строке в первом столбце и получаем номер CVE
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
    cve_number = row[0]
    vulnerability_info = get_vulnerability_info(cve_number)
    print(vulnerability_info)

# Сохраняем изменения в Excel-файл
workbook.save('output.xlsx')
